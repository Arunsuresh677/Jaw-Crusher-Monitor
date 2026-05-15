"""
reports.py — PDF, CSV and Excel report generation
Crusher Monitor — Kannan Blue Metals

All reports query the database for the requested date range.
Called by GET /api/reports/pdf, /api/reports/csv, /api/reports/excel
"""

import csv
import io
import logging
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

log = logging.getLogger(__name__)


# ── Time helpers ─────────────────────────────────────────────────────────

def _to_12h(dt_str: str) -> str:
    """Convert 'YYYY-MM-DD HH:MM:SS' or 'HH:MM:SS' to 12-hour format."""
    if not dt_str or dt_str in ("—", "None", "null"):
        return "—"
    try:
        s = dt_str[:19]
        fmt = "%Y-%m-%d %H:%M:%S" if len(s) == 19 and "-" in s else "%H:%M:%S"
        dt = datetime.strptime(s, fmt)
        return dt.strftime("%I:%M:%S %p").lstrip("0") or "12:00:00 AM"
    except Exception:
        return dt_str


def _hms_to_secs(hms: str) -> int:
    """Convert HH:MM:SS duration string to total seconds."""
    try:
        h, m, s = hms.split(":")
        return int(h) * 3600 + int(m) * 60 + int(s)
    except Exception:
        return 0


def _secs_to_hms(secs: int) -> str:
    """Convert total seconds to HH:MM:SS duration string."""
    secs = max(0, int(secs))
    return f"{secs // 3600:02d}:{(secs % 3600) // 60:02d}:{secs % 60:02d}"


def _calc_elapsed(start_str: str, end_str: str) -> str:
    """Calculate End − Start and return as HH:MM:SS."""
    try:
        fmt   = "%Y-%m-%d %H:%M:%S"
        start = datetime.strptime(start_str[:19], fmt)
        end   = datetime.strptime(end_str[:19],   fmt)
        return _secs_to_hms(int((end - start).total_seconds()))
    except Exception:
        return "—"


def _fmt_date_range(from_date: str, to_date: str) -> str:
    """Format date range: '15 May 2026' or '10 May – 15 May 2026'"""
    try:
        fmt = "%Y-%m-%d"
        fd  = datetime.strptime(from_date, fmt).strftime("%d %b %Y")
        td  = datetime.strptime(to_date,   fmt).strftime("%d %b %Y")
        return fd if fd == td else f"{fd} – {td}"
    except Exception:
        return f"{from_date} – {to_date}"


def _build_metrics(summary: dict) -> dict:
    """Derive all display values from the DB summary dict."""
    run_secs     = int(summary.get("run_secs",     0) or 0)
    stuck_secs   = int(summary.get("stuck_secs",   0) or 0)
    no_feed_secs = int(summary.get("no_feed_secs", 0) or 0)
    prod_secs    = max(0, run_secs - stuck_secs - no_feed_secs)
    avail_pct    = float(summary.get("availability_pct", 0) or 0)

    first_run_at = summary.get("first_run_at")
    last_seen    = summary.get("last_seen")

    # End time — show "Running" if seen within last 2 minutes
    end_disp = "—"
    if last_seen:
        try:
            last_dt = datetime.strptime(last_seen[:19], "%Y-%m-%d %H:%M:%S")
            if (datetime.now() - last_dt).total_seconds() < 120:
                end_disp = f"Running  ({_to_12h(last_seen)})"
            else:
                end_disp = _to_12h(last_seen)
        except Exception:
            end_disp = _to_12h(last_seen)

    # Total Run Time = End − Production Start (wall clock)
    if first_run_at and last_seen:
        total_run_hms = _calc_elapsed(first_run_at, last_seen)
    else:
        total_run_hms = _secs_to_hms(run_secs) if run_secs else "—"

    return {
        "prod_start":   _to_12h(first_run_at) if first_run_at else "Not yet started",
        "end_time":     end_disp,
        "total_run":    total_run_hms,
        "stuck_hms":    _secs_to_hms(stuck_secs),
        "no_feed_hms":  _secs_to_hms(no_feed_secs),
        "total_prod":   _secs_to_hms(prod_secs) if run_secs else "—",
        "avail_pct":    avail_pct,
        "total_alerts": int(summary.get("total_alerts", 0) or 0),
        "peak_vfd":     summary.get("peak_vfd_hz", 0) or 0,
        "avg_vfd":      summary.get("avg_vfd_hz",  0) or 0,
    }


# ── Brand colours ────────────────────────────────────────────────────────
AMBER  = colors.HexColor("#F5A623")
NAVY   = colors.HexColor("#1A2B5E")
GREEN  = colors.HexColor("#1DB97A")
RED    = colors.HexColor("#E03A3A")
ORANGE = colors.HexColor("#F07C20")
GREY   = colors.HexColor("#8892B0")
LIGHT  = colors.HexColor("#F0F2FF")
WHITE  = colors.white
BLACK  = colors.HexColor("#0D1320")
BLUE   = colors.HexColor("#1A6FD4")


def _avail_color(val: float) -> colors.HexColor:
    if val >= 85: return GREEN
    if val >= 65: return ORANGE
    return RED


def _alert_color(n: int) -> colors.HexColor:
    if n > 50: return RED
    if n > 10: return ORANGE
    return GREEN


# ── Style helpers ────────────────────────────────────────────────────────
def _styles():
    return {
        "title": ParagraphStyle(
            "title", fontName="Helvetica-Bold", fontSize=22,
            leading=26, textColor=NAVY, spaceAfter=0, alignment=TA_LEFT,
        ),
        "subtitle": ParagraphStyle(
            "subtitle", fontName="Helvetica", fontSize=9.5,
            leading=14, textColor=GREY, spaceAfter=0, alignment=TA_LEFT,
        ),
        "section": ParagraphStyle(
            "section", fontName="Helvetica-Bold", fontSize=11,
            textColor=NAVY, spaceBefore=14, spaceAfter=6,
        ),
        "small": ParagraphStyle(
            "small", fontName="Helvetica", fontSize=8,
            textColor=GREY, alignment=TA_CENTER,
        ),
        "meta": ParagraphStyle(
            "meta", fontName="Helvetica", fontSize=9,
            textColor=BLACK, spaceAfter=3,
        ),
    }


# ── PDF builder ──────────────────────────────────────────────────────────

def build_period_pdf(summary: dict, period_label: str,
                     from_date: str, to_date: str) -> bytes:
    """
    Build a PDF report from DB period summary.
    summary comes from database.get_period_summary().
    Returns raw PDF bytes.
    """
    buf        = io.BytesIO()
    generated  = datetime.now().strftime("%d %b %Y  %H:%M")
    date_range = _fmt_date_range(from_date, to_date)

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=18*mm, bottomMargin=16*mm,
        title="Crusher Monitor Report — Kannan Blue Metals",
        author="Crusher Monitor",
    )
    S     = _styles()
    story = []

    # ── Header ───────────────────────────────────────────────
    story.append(Paragraph("Kannan Blue Metals", S["title"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"Jaw Crusher Monitor — {period_label}  ·  {date_range}  ·  "
        f"Chennimalai, Erode  ·  Generated {generated}",
        S["subtitle"],
    ))
    story.append(Spacer(1, 14))

    div = Table([[""]], colWidths=[doc.width], rowHeights=[1.5])
    div.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), AMBER),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ("TOPPADDING",    (0,0),(-1,-1), 0),
        ("BOTTOMPADDING", (0,0),(-1,-1), 0),
    ]))
    story.append(div)
    story.append(Spacer(1, 16))

    # ── Period Summary table ──────────────────────────────────
    story.append(Paragraph(f"Period Summary  ·  {period_label}", S["section"]))

    m     = _build_metrics(summary)
    avail = m["avail_pct"]
    nalerts = m["total_alerts"]

    table_data = [
        ["Metric",                  "Value"],
        ["Production Start Time",   m["prod_start"]],
        ["End Time",                m["end_time"]],
        ["Total Run Time",          m["total_run"]],
        ["Stone Stuck Time",        m["stuck_hms"]],
        ["No Raw Material Time",    m["no_feed_hms"]],
        ["Total Production Time",   m["total_prod"]],
        ["Availability",            f"{avail:.1f} %"],
        ["Total Alerts",            str(nalerts)],
        ["Peak VFD",                f"{m['peak_vfd']} Hz"],
        ["Avg VFD",                 f"{m['avg_vfd']} Hz"],
    ]

    col_w = [80*mm, 80*mm]
    t = Table(table_data, colWidths=col_w)
    t.setStyle(TableStyle([
        # Header
        ("BACKGROUND",    (0,0),(-1,0), NAVY),
        ("TEXTCOLOR",     (0,0),(-1,0), WHITE),
        ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0), 10),
        ("ALIGN",         (0,0),(-1,0), "LEFT"),
        # Body
        ("BACKGROUND",    (0,1),(0,-1), LIGHT),
        ("FONTNAME",      (0,1),(0,-1), "Helvetica-Bold"),
        ("FONTNAME",      (1,1),(1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1),(-1,-1), 10),
        ("GRID",          (0,0),(-1,-1), 0.4, colors.HexColor("#D0D4E8")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [WHITE, colors.HexColor("#F7F8FC")]),
        ("LEFTPADDING",   (0,0),(-1,-1), 8),
        ("RIGHTPADDING",  (0,0),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 7),
        ("BOTTOMPADDING", (0,0),(-1,-1), 7),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        # Row 1 — Production Start — blue
        ("TEXTCOLOR",     (1,1),(1,1), BLUE),
        ("FONTNAME",      (1,1),(1,1), "Helvetica-Bold"),
        # Row 2 — End Time — grey
        ("TEXTCOLOR",     (1,2),(1,2), GREY),
        # Row 3 — Total Run — green
        ("TEXTCOLOR",     (1,3),(1,3), GREEN),
        ("FONTNAME",      (1,3),(1,3), "Helvetica-Bold"),
        # Row 4 — Stone Stuck — red
        ("TEXTCOLOR",     (1,4),(1,4), RED),
        ("FONTNAME",      (1,4),(1,4), "Helvetica-Bold"),
        # Row 5 — No Raw Material — orange
        ("TEXTCOLOR",     (1,5),(1,5), ORANGE),
        ("FONTNAME",      (1,5),(1,5), "Helvetica-Bold"),
        # Row 6 — Total Production — green + highlight
        ("TEXTCOLOR",     (1,6),(1,6), GREEN),
        ("FONTNAME",      (1,6),(1,6), "Helvetica-Bold"),
        ("BACKGROUND",    (0,6),(-1,6), colors.HexColor("#EAF7F1")),
        # Row 7 — Availability — dynamic
        ("TEXTCOLOR",     (1,7),(1,7), _avail_color(avail)),
        ("FONTNAME",      (1,7),(1,7), "Helvetica-Bold"),
        # Row 8 — Total Alerts — dynamic
        ("TEXTCOLOR",     (1,8),(1,8), _alert_color(nalerts)),
        # Row 9 — Peak VFD — navy
        ("TEXTCOLOR",     (1,9),(1,9), NAVY),
        ("FONTNAME",      (1,9),(1,9), "Helvetica-Bold"),
        # Row 10 — Avg VFD — grey
        ("TEXTCOLOR",     (1,10),(1,10), GREY),
    ]))
    story.append(t)
    story.append(Spacer(1, 20))

    # ── Footer ───────────────────────────────────────────────
    fdiv = Table([[""]], colWidths=[doc.width], rowHeights=[0.8])
    fdiv.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), GREY),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ("TOPPADDING",    (0,0),(-1,-1), 0),
        ("BOTTOMPADDING", (0,0),(-1,-1), 0),
    ]))
    story.append(fdiv)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"Crusher Monitor  ·  Kannan Blue Metals, Chennimalai, Erode  ·  "
        f"Report generated {generated}",
        S["small"],
    ))

    doc.build(story)
    return buf.getvalue()


# ── CSV builder ──────────────────────────────────────────────────────────

def build_period_csv(summary: dict, period_label: str,
                     from_date: str, to_date: str) -> str:
    """Build CSV from DB period summary. Returns UTF-8 string."""
    buf        = io.StringIO()
    date_range = _fmt_date_range(from_date, to_date)
    m          = _build_metrics(summary)

    buf.write("# Kannan Blue Metals — Crusher Monitor Report\n")
    buf.write(f"# Period: {period_label}  ({date_range})\n")
    buf.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    buf.write("#\n")
    buf.write("metric,value\n")
    buf.write(f"Period,{period_label}\n")
    buf.write(f"Date Range,{date_range}\n")
    buf.write(f"Production Start Time,{m['prod_start']}\n")
    buf.write(f"End Time,{m['end_time']}\n")
    buf.write(f"Total Run Time,{m['total_run']}\n")
    buf.write(f"Stone Stuck Time,{m['stuck_hms']}\n")
    buf.write(f"No Raw Material Time,{m['no_feed_hms']}\n")
    buf.write(f"Total Production Time,{m['total_prod']}\n")
    buf.write(f"Availability %,{m['avail_pct']:.1f}\n")
    buf.write(f"Total Alerts,{m['total_alerts']}\n")
    buf.write(f"Peak VFD Hz,{m['peak_vfd']}\n")
    buf.write(f"Avg VFD Hz,{m['avg_vfd']}\n")
    return buf.getvalue()


# ── Excel builder ────────────────────────────────────────────────────────

def build_period_excel(summary: dict, period_label: str,
                       from_date: str, to_date: str) -> bytes:
    """
    Build an Excel .xlsx from DB period summary.
    Requires: pip install openpyxl
    Returns raw bytes.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Period Summary"

    date_range = _fmt_date_range(from_date, to_date)
    m          = _build_metrics(summary)
    avail      = m["avail_pct"]
    nalerts    = m["total_alerts"]
    generated  = datetime.now().strftime("%d %b %Y  %H:%M")

    # ── Fill styles ───────────────────────────────────────────
    navy_fill  = PatternFill("solid", fgColor="1A2B5E")
    green_fill = PatternFill("solid", fgColor="EAF7F1")
    light_fill = PatternFill("solid", fgColor="F0F2FF")
    alt_fill   = PatternFill("solid", fgColor="F7F8FC")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    # ── Border ────────────────────────────────────────────────
    thin   = Side(style="thin", color="D0D4E8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left   = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _cell(row, col, value, font_color="0D1320", bold=False,
              fill=white_fill, size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Calibri", bold=bold, size=size,
                           color=font_color)
        c.fill      = fill
        c.alignment = left
        c.border    = border
        return c

    # ── Title block (no border) ───────────────────────────────
    ws.merge_cells("A1:B1")
    c = ws.cell(row=1, column=1, value="Kannan Blue Metals — Crusher Monitor")
    c.font      = Font(name="Calibri", bold=True, size=14, color="1A2B5E")
    c.alignment = left

    ws.merge_cells("A2:B2")
    c = ws.cell(row=2, column=1,
                value=f"{period_label}  ·  {date_range}  ·  Chennimalai, Erode")
    c.font      = Font(name="Calibri", size=9, color="8892B0")
    c.alignment = left

    ws.merge_cells("A3:B3")
    c = ws.cell(row=3, column=1, value=f"Generated: {generated}")
    c.font      = Font(name="Calibri", size=9, color="8892B0")
    c.alignment = left

    # Row 4 — blank spacer
    ws.append([])

    # ── Header row (row 5) ────────────────────────────────────
    hdr_row = 5
    for col, label in enumerate(["Metric", "Value"], start=1):
        c = ws.cell(row=hdr_row, column=col, value=label)
        c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill      = navy_fill
        c.alignment = left
        c.border    = border

    # ── Data rows ─────────────────────────────────────────────
    # (metric, value, value_color_hex, value_bold, row_fill)
    avail_hex  = "1DB97A" if avail >= 85 else "F07C20" if avail >= 65 else "E03A3A"
    alert_hex  = "E03A3A" if nalerts > 50 else "F07C20" if nalerts > 10 else "1DB97A"

    data_rows = [
        ("Production Start Time",  m["prod_start"],        "1A6FD4", True,  white_fill),
        ("End Time",               m["end_time"],           "8892B0", False, alt_fill),
        ("Total Run Time",         m["total_run"],          "1DB97A", True,  white_fill),
        ("Stone Stuck Time",       m["stuck_hms"],          "E03A3A", True,  alt_fill),
        ("No Raw Material Time",   m["no_feed_hms"],        "F07C20", True,  white_fill),
        ("Total Production Time",  m["total_prod"],         "1DB97A", True,  green_fill),
        ("Availability",           f"{avail:.1f} %",        avail_hex,True,  alt_fill),
        ("Total Alerts",           str(nalerts),            alert_hex,False, white_fill),
        ("Peak VFD",               f"{m['peak_vfd']} Hz",  "1A2B5E", True,  alt_fill),
        ("Avg VFD",                f"{m['avg_vfd']} Hz",   "8892B0", False, white_fill),
    ]

    for i, (metric, value, vcolor, vbold, rfill) in enumerate(data_rows):
        r = hdr_row + 1 + i
        _cell(r, 1, metric, "0D1320", bold=True,  fill=light_fill)
        _cell(r, 2, value,  vcolor,   bold=vbold,  fill=rfill)

    # ── Column widths / row heights ───────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 26
    ws.row_dimensions[1].height     = 22
    ws.row_dimensions[2].height     = 14
    ws.row_dimensions[3].height     = 14
    for r in range(hdr_row, hdr_row + 1 + len(data_rows) + 1):
        ws.row_dimensions[r].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Backward-compat aliases (old main.py endpoints) ──────────────────────

def build_shift_pdf(shift_rows: list, live_state: dict) -> bytes:
    """Legacy wrapper — redirects to DB-based builder using live state."""
    from datetime import date as _date
    today   = _date.today().strftime("%Y-%m-%d")
    summary = {
        "run_secs":         _hms_to_secs(live_state.get("timer_run",     "00:00:00")),
        "stuck_secs":       _hms_to_secs(live_state.get("timer_stuck",   "00:00:00")),
        "no_feed_secs":     _hms_to_secs(live_state.get("timer_no_feed", "00:00:00")),
        "first_run_at":     live_state.get("first_run_at"),
        "last_seen":        live_state.get("last_active"),
        "availability_pct": live_state.get("availability_pct", 0),
        "total_alerts":     live_state.get("alert_count", 0),
        "peak_vfd_hz":      0,
        "avg_vfd_hz":       0,
    }
    return build_period_pdf(summary, "Daily Report", today, today)


def build_shift_csv(shift_rows: list, live_state: dict | None = None) -> str:
    """Legacy wrapper — redirects to DB-based builder using live state."""
    from datetime import date as _date
    today   = _date.today().strftime("%Y-%m-%d")
    summary = {}
    if live_state:
        summary = {
            "run_secs":         _hms_to_secs(live_state.get("timer_run",     "00:00:00")),
            "stuck_secs":       _hms_to_secs(live_state.get("timer_stuck",   "00:00:00")),
            "no_feed_secs":     _hms_to_secs(live_state.get("timer_no_feed", "00:00:00")),
            "first_run_at":     live_state.get("first_run_at"),
            "last_seen":        live_state.get("last_active"),
            "availability_pct": live_state.get("availability_pct", 0),
            "total_alerts":     live_state.get("alert_count", 0),
            "peak_vfd_hz":      0,
            "avg_vfd_hz":       0,
        }
    return build_period_csv(summary, "Daily Report", today, today)
